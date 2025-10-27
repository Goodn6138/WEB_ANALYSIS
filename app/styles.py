from openpyxl.styles import Font, PatternFill, Border, Side

header_fill = PatternFill("solid", fgColor="003366FF")
odd_fill    = PatternFill("solid", fgColor="00E7F3FF")
font_white  = Font(color="FFFFFF", bold=True)
font_bold   = Font(bold=True)
thin_border = Border(left=Side(style='thin'), right=Side(style='thin'),
                     top=Side(style='thin'), bottom=Side(style='thin'))
