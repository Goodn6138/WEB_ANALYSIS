from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Alignment
from app.styles import *

def write_table(ws, df, start_row, start_col, title):
    end_col = start_col + len(df.columns) - 1
    ws.merge_cells(start_row=start_row, start_column=start_col, end_row=start_row, end_column=end_col)
    title_cell = ws.cell(row=start_row, column=start_col, value=title)
    title_cell.font = font_bold
    title_cell.alignment = Alignment(horizontal="center")

    start_row += 1
    for c, col_name in enumerate(df.columns, start=start_col):
        cell = ws.cell(row=start_row, column=c, value=col_name)
        cell.fill, cell.font, cell.border = header_fill, font_white, thin_border

    for r_off, row in enumerate(dataframe_to_rows(df, index=False, header=False), start=1):
        for c_off, val in enumerate(row, start=start_col):
            cell = ws.cell(row=start_row + r_off, column=c_off, value=val)
            cell.border = thin_border
            if r_off % 2 == 0:
                cell.fill = odd_fill

    return start_row + len(df) + 2
